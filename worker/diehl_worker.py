import json, os, socket, subprocess, sys, time
from pathlib import Path
import requests
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parent

def load_env():
 p=ROOT/'worker.env'
 if p.exists():
  for line in p.read_text(encoding='utf-8').splitlines():
   if '=' in line and not line.lstrip().startswith('#'):
    k,v=line.split('=',1);os.environ.setdefault(k.strip(),os.path.expandvars(v.strip()))
load_env()
BASE=os.environ.get('PLATFORM_URL','').rstrip('/');SECRET=os.environ.get('DTNA_WORKER_SECRET','');WORKBOOK=Path(os.path.expandvars(os.environ.get('MASTER_WORKBOOK','')));POLL=float(os.environ.get('POLL_SECONDS','5'));WORKER=f"{socket.gethostname()}-{os.getpid()}"
HEAD={'x-worker-secret':SECRET,'content-type':'application/json'}

def post(path,payload):
 r=requests.post(BASE+path,headers=HEAD,json=payload,timeout=90);r.raise_for_status();return r.json()
def proc_running(name):
 try:
  import psutil
  return any(name.lower() in (p.name() or '').lower() for p in psutil.process_iter())
 except Exception:return False

def heartbeat(detail=None):
 post('/api/worker/heartbeat',{'workerId':WORKER,'hostname':socket.gethostname(),'dtnaStatus':'browser ready' if proc_running('msedge') or proc_running('chrome') else 'worker ready','outlookStatus':'open' if proc_running('outlook') else 'available','onedriveStatus':'connected' if WORKBOOK.exists() else 'workbook missing','masterWorkbook':str(WORKBOOK),'details':detail or {}})

def read_master(vins):
 out={}
 if not WORKBOOK.exists():return out
 wb=load_workbook(WORKBOOK,read_only=True,data_only=True);ws=wb['VIN Data'] if 'VIN Data' in wb.sheetnames else wb.active
 headers=[str(c.value or '').strip() for c in next(ws.iter_rows())];idx={h:i for i,h in enumerate(headers)};want=set(vins)
 for row in ws.iter_rows(values_only=True):
  vin=str(row[idx.get('VIN',0)] or '').strip().upper()
  if vin in want:
   data={headers[i]:row[i] for i in range(min(len(headers),len(row))) if headers[i] and row[i] not in (None,'')}
   out[vin]=data
   if len(out)==len(want):break
 wb.close();return out

def normalize(vin,row):
 def g(*names):
  for n in names:
   if n in row and row[n] not in (None,''):return row[n]
  return None
 val=lambda x:x.isoformat() if hasattr(x,'isoformat') else x
 return {'vin':vin,'verificationStatus':g('Verification Status'),'inServiceStatus':g('In-Service Status'),'inServiceDate':val(g('In-Service Date')),'mileage':g('Mileage'),'customerResult':g('Customer Result'),'customerName':g('Customer Name'),'registeredCustomerName':g('Registered Customer Name'),'registeredCustomerAccount':g('Registered Customer Account'),'orderedCustomerName':g('Ordered Customer Name'),'raw':{k:val(v) for k,v in row.items()}}

def optional_lookup(vins):
 cmd=os.environ.get('VIN_LOOKUP_COMMAND','').strip()
 if not cmd:return
 env=os.environ.copy();env['DIEHL_VINS']='\n'.join(vins)
 subprocess.run(cmd,shell=True,env=env,check=False)

def run_dtna_sync_once():
 script=os.environ.get('DTNA_SYNC_SCRIPT','').strip()
 if not script:return
 p=Path(os.path.expandvars(script));p=p if p.is_absolute() else (ROOT/p).resolve()
 if p.exists():subprocess.run([sys.executable,str(p)],cwd=str(p.parent),check=False)

def process(claim):
 batch=claim.get('batch');items=claim.get('items') or []
 if not batch or not items:return False
 vins=[x['vin'] for x in items];heartbeat({'batchId':batch['id'],'stage':'starting','count':len(vins)})
 # The existing DTNA script opens the persistent browser profile and handles MFA when needed.
 run_dtna_sync_once();found=read_master(vins)
 missing=[v for v in vins if v not in found]
 if missing:optional_lookup(missing);found.update(read_master(missing))
 for item in items:
  vin=item['vin'];row=found.get(vin)
  if row:
   post('/api/worker/result',{'batchId':batch['id'],'itemId':item['id'],'status':'complete','attempts':int(item.get('attempts') or 0)+1,'result':normalize(vin,row)})
  else:
   post('/api/worker/result',{'batchId':batch['id'],'itemId':item['id'],'status':'error','attempts':int(item.get('attempts') or 0)+1,'error':'VIN not found in master workbook after local lookup/sync'})
 heartbeat({'batchId':batch['id'],'stage':'idle'});return True

def main():
 if not BASE or not SECRET:raise SystemExit('Set PLATFORM_URL and DTNA_WORKER_SECRET in worker.env')
 print('Diehl worker online:',WORKER);print('Platform:',BASE);print('Workbook:',WORKBOOK)
 while True:
  try:
   heartbeat();claim=post('/api/worker/claim',{'workerId':WORKER,'limit':20});process(claim)
  except KeyboardInterrupt:return
  except Exception as e:print('worker error:',e)
  time.sleep(POLL)
if __name__=='__main__':main()
