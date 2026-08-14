from __future__ import annotations

import json
import os
import re
import socket
import sqlite3
import subprocess
import threading
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel

ROOT = Path(__file__).resolve().parent
CONFIG = ROOT / 'config.json'
DB = ROOT / 'worker_state.db'
DTNA_SCRIPT = ROOT / 'dtna_login_and_sync.py'
PROFILE = Path(os.environ.get('LOCALAPPDATA', str(ROOT))) / 'DiehlDTNAManual' / 'browser_profile'
VERSION = '4.0'
PORT = 8765

app = FastAPI(title='Diehl VIN Worker', version=VERSION)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        'https://diehl-vin-platform.vercel.app',
        'http://localhost:3000',
        'http://127.0.0.1:3000',
    ],
    allow_credentials=False,
    allow_methods=['*'],
    allow_headers=['*'],
)


@app.middleware('http')
async def private_network_access(request: Request, call_next):
    response = await call_next(request)
    if request.headers.get('access-control-request-private-network') == 'true':
        response.headers['Access-Control-Allow-Private-Network'] = 'true'
    response.headers['Cache-Control'] = 'no-store'
    return response


excel_lock = threading.RLock()
stop_event = threading.Event()


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def config() -> dict[str, Any]:
    try:
        data = json.loads(CONFIG.read_text(encoding='utf-8')) if CONFIG.exists() else {}
    except Exception:
        data = {}
    data.setdefault('masterWorkbook', '')
    data.setdefault('port', PORT)
    data.setdefault('vinLookupCommand', f'"{ROOT / ".venv" / "Scripts" / "python.exe"}" "{ROOT / "vin_lookup.py"}"')
    return data


def workbook_path() -> Path:
    return Path(os.path.expandvars(str(config().get('masterWorkbook') or ''))).expanduser()


def conn() -> sqlite3.Connection:
    c = sqlite3.connect(DB, check_same_thread=False, timeout=30)
    c.row_factory = sqlite3.Row
    return c


def init_db() -> None:
    c = conn()
    c.executescript('''
    create table if not exists batches(
      id text primary key,
      status text not null,
      total_vins integer not null,
      lookup_mode text not null,
      options text not null,
      created_at text not null,
      started_at text,
      completed_at text
    );
    create table if not exists items(
      id text primary key,
      batch_id text not null,
      vin text not null,
      queue_position integer not null,
      status text not null,
      attempts integer not null default 0,
      result text,
      error_message text,
      started_at text,
      completed_at text
    );
    create index if not exists items_batch_idx on items(batch_id, queue_position);
    ''')
    c.commit(); c.close()


init_db()


def serialize(v: Any) -> Any:
    return v.isoformat() if hasattr(v, 'isoformat') else v


def find_open_workbook(excel: Any, path: Path):
    target = os.path.normcase(os.path.abspath(str(path)))
    try:
        for wb in excel.Workbooks:
            try:
                if os.path.normcase(os.path.abspath(str(wb.FullName))) == target:
                    return wb
            except Exception:
                pass
    except Exception:
        pass
    return None


def read_master_com(path: Path, wanted: set[str]) -> dict[str, dict[str, Any]]:
    import pythoncom
    import win32com.client
    pythoncom.CoInitialize()
    excel = wb = None
    opened_here = created_excel = False
    try:
        try:
            excel = win32com.client.GetActiveObject('Excel.Application')
        except Exception:
            excel = win32com.client.DispatchEx('Excel.Application')
            excel.Visible = False; excel.DisplayAlerts = False; created_excel = True
        wb = find_open_workbook(excel, path)
        if wb is None:
            wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=True, IgnoreReadOnlyRecommended=True)
            opened_here = True
        try:
            ws = wb.Worksheets('VIN Data')
        except Exception:
            ws = wb.Worksheets(1)
        cols = max(1, int(ws.UsedRange.Columns.Count))
        headers = [str(ws.Cells(1, i).Value or '').strip() for i in range(1, cols + 1)]
        if 'VIN' not in headers:
            raise RuntimeError('Workbook must contain a VIN column in row 1.')
        vin_col = headers.index('VIN') + 1
        last = int(ws.Cells(ws.Rows.Count, vin_col).End(-4162).Row)
        out: dict[str, dict[str, Any]] = {}
        for r in range(2, max(2, last) + 1):
            vin = str(ws.Cells(r, vin_col).Value or '').strip().upper()
            if vin in wanted:
                row = {}
                for c, header in enumerate(headers, start=1):
                    if not header: continue
                    val = ws.Cells(r, c).Value
                    if val not in (None, ''): row[header] = serialize(val)
                out[vin] = row
                if len(out) == len(wanted): break
        return out
    finally:
        try:
            if wb is not None and opened_here: wb.Close(SaveChanges=False)
        except Exception: pass
        try:
            if excel is not None and created_excel: excel.Quit()
        except Exception: pass
        pythoncom.CoUninitialize()


def read_master_openpyxl(path: Path, wanted: set[str]) -> dict[str, dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=path.suffix.lower() == '.xlsm')
    try:
        ws = wb['VIN Data'] if 'VIN Data' in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(x or '').strip() for x in next(rows)]
        if 'VIN' not in headers:
            raise RuntimeError('Workbook must contain a VIN column in row 1.')
        vi = headers.index('VIN')
        out = {}
        for row in rows:
            vin = str(row[vi] or '').strip().upper()
            if vin in wanted:
                out[vin] = {headers[i]: serialize(row[i]) for i in range(min(len(headers), len(row))) if headers[i] and row[i] not in (None, '')}
                if len(out) == len(wanted): break
        return out
    finally:
        wb.close()


def read_master(wanted: set[str]) -> dict[str, dict[str, Any]]:
    path = workbook_path()
    if not path.exists():
        raise RuntimeError('The configured Excel workbook cannot be found. Choose it again on the Initializer page.')
    with excel_lock:
        com_error = None
        try:
            return read_master_com(path, wanted)
        except Exception as exc:
            com_error = exc
        try:
            return read_master_openpyxl(path, wanted)
        except Exception as exc:
            raise RuntimeError(f'Could not read the workbook. Excel: {com_error}; file: {exc}') from exc


def normalize(vin: str, row: dict[str, Any]) -> dict[str, Any]:
    def get(*names: str):
        for n in names:
            if row.get(n) not in (None, ''): return row[n]
        return None
    return {
        'vin': vin,
        'verificationStatus': get('Verification Status'),
        'inServiceStatus': get('In-Service Status'),
        'inServiceDate': get('In-Service Date'),
        'mileage': get('Mileage'),
        'customerResult': get('Customer Result'),
        'customerName': get('Customer Name'),
        'registeredCustomerName': get('Registered Customer Name'),
        'registeredCustomerAccount': get('Registered Customer Account'),
        'orderedCustomerName': get('Ordered Customer Name'),
        'raw': row,
    }


def write_result(vin: str, result: dict[str, Any]) -> None:
    import pythoncom
    import win32com.client
    path = workbook_path()
    mapping = {
        'verificationStatus': 'Verification Status',
        'inServiceStatus': 'In-Service Status',
        'inServiceDate': 'In-Service Date',
        'mileage': 'Mileage',
        'customerResult': 'Customer Result',
        'customerName': 'Customer Name',
        'registeredCustomerName': 'Registered Customer Name',
        'registeredCustomerAccount': 'Registered Customer Account',
        'orderedCustomerName': 'Ordered Customer Name',
    }
    with excel_lock:
        pythoncom.CoInitialize()
        excel = wb = None
        opened_here = created_excel = False
        try:
            try:
                excel = win32com.client.GetActiveObject('Excel.Application')
            except Exception:
                excel = win32com.client.DispatchEx('Excel.Application')
                excel.Visible = False; excel.DisplayAlerts = False; created_excel = True
            wb = find_open_workbook(excel, path)
            if wb is None:
                wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=False, IgnoreReadOnlyRecommended=True)
                opened_here = True
            if bool(getattr(wb, 'ReadOnly', False)):
                raise RuntimeError('The workbook is currently read-only. Close duplicate copies and allow OneDrive to finish syncing.')
            try:
                ws = wb.Worksheets('VIN Data')
            except Exception:
                ws = wb.Worksheets(1)
            cols = max(1, int(ws.UsedRange.Columns.Count))
            headers = {}
            for c in range(1, cols + 1):
                h = str(ws.Cells(1, c).Value or '').strip()
                if h: headers[h] = c
            if 'VIN' not in headers: raise RuntimeError('Workbook must contain a VIN column in row 1.')
            for h in mapping.values():
                if h not in headers:
                    cols += 1; ws.Cells(1, cols).Value = h; headers[h] = cols
            vin_col = headers['VIN']
            last = int(ws.Cells(ws.Rows.Count, vin_col).End(-4162).Row)
            target = None
            for r in range(2, max(2, last) + 1):
                if str(ws.Cells(r, vin_col).Value or '').strip().upper() == vin:
                    target = r; break
            if target is None:
                target = max(2, last + 1)
                if target > 2:
                    ws.Rows(target - 1).Copy()
                    ws.Rows(target).PasteSpecial(Paste=-4122)
                    ws.Rows(target).PasteSpecial(Paste=6)
                    excel.CutCopyMode = False
                ws.Cells(target, vin_col).Value = vin
            for key, header in mapping.items():
                val = result.get(key)
                if val not in (None, ''): ws.Cells(target, headers[header]).Value = val
            wb.Save()
        finally:
            try:
                if wb is not None and opened_here: wb.Close(SaveChanges=True)
            except Exception: pass
            try:
                if excel is not None and created_excel: excel.Quit()
            except Exception: pass
            pythoncom.CoUninitialize()


def run_lookup(vins: list[str]) -> dict[str, Any]:
    if not vins: return {}
    result_file = ROOT / 'vin-results-v4.json'
    result_file.unlink(missing_ok=True)
    env = os.environ.copy()
    env['DIEHL_VINS'] = '\n'.join(vins)
    env['DIEHL_RESULT_FILE'] = str(result_file)
    py = ROOT / '.venv' / 'Scripts' / 'python.exe'
    completed = subprocess.run([str(py), str(ROOT / 'vin_lookup.py')], cwd=str(ROOT), env=env, check=False)
    if completed.returncode != 0 or not result_file.exists():
        return {'_error': 'DTNA lookup failed. Check the DTNA browser/log window.'}
    try:
        data = json.loads(result_file.read_text(encoding='utf-8'))
        return data if isinstance(data, dict) else {}
    except Exception as exc:
        return {'_error': str(exc)}


def scheduler() -> None:
    while not stop_event.is_set():
        c = None
        try:
            c = conn()
            batch = c.execute("select * from batches where status in ('queued','running') order by created_at limit 1").fetchone()
            if not batch:
                c.close(); c = None; time.sleep(.5); continue
            c.execute("update batches set status='running', started_at=coalesce(started_at, ?) where id=?", (now(), batch['id']))
            items = c.execute("select * from items where batch_id=? and status in ('queued','retry') order by queue_position", (batch['id'],)).fetchall()
            if not items:
                remaining = c.execute("select count(*) n from items where batch_id=? and status='running'", (batch['id'],)).fetchone()['n']
                if remaining == 0: c.execute("update batches set status='complete', completed_at=? where id=?", (now(), batch['id']))
                c.commit(); c.close(); c = None; time.sleep(.3); continue
            for item in items:
                c.execute("update items set status='running', started_at=?, attempts=attempts+1 where id=?", (now(), item['id']))
            c.commit(); c.close(); c = None

            wanted = {str(x['vin']).upper() for x in items}
            try:
                existing = read_master(wanted)
            except Exception as exc:
                x = conn()
                for item in items:
                    x.execute("update items set status='error', error_message=?, completed_at=? where id=?", (str(exc), now(), item['id']))
                x.commit(); x.close(); continue

            missing = [vin for vin in wanted if vin not in existing]
            looked = run_lookup(missing) if missing else {}
            lookup_error = looked.get('_error') if isinstance(looked, dict) else None

            x = conn()
            for item in items:
                vin = str(item['vin']).upper()
                try:
                    if vin in existing:
                        result = normalize(vin, existing[vin])
                    else:
                        result = looked.get(vin) if isinstance(looked, dict) else None
                        if not result:
                            raise RuntimeError(lookup_error or 'VIN was not found in the workbook or DTNA cache.')
                        result.setdefault('vin', vin)
                        write_result(vin, result)
                    x.execute("update items set status='complete', result=?, error_message=null, completed_at=? where id=?", (json.dumps(result, default=str), now(), item['id']))
                except Exception as exc:
                    x.execute("update items set status='error', error_message=?, completed_at=? where id=?", (str(exc), now(), item['id']))
            x.commit(); x.close()
        except Exception as exc:
            print('scheduler error:', exc, flush=True)
            time.sleep(1)
        finally:
            try:
                if c is not None: c.close()
            except Exception: pass


threading.Thread(target=scheduler, daemon=True).start()


class BatchIn(BaseModel):
    vins: Any
    lookupMode: str = 'in_service_customer'
    workers: int = 1
    batchSize: int = 100
    retryRounds: int = 3
    batchPause: float = .5
    retryPause: float = 3


def clean_vins(value: Any) -> list[str]:
    raw = '\n'.join(value) if isinstance(value, list) else str(value or '')
    return list(dict.fromkeys(x.upper() for x in re.split(r'[\s,;]+', raw) if re.fullmatch(r'[A-HJ-NPR-Z0-9]{17}', x.upper())))


@app.get('/ping')
def ping():
    return {'ok': True, 'product': 'DiehlVINWorker', 'version': VERSION, 'hostname': socket.gethostname()}


@app.get('/health')
def health():
    path = workbook_path()
    return {
        'ok': True,
        'version': VERSION,
        'worker': {
            'worker_id': socket.gethostname(),
            'hostname': socket.gethostname(),
            'dtna_status': 'ready',
            'master_workbook': str(path) if str(path) else '',
            'last_seen': now(),
            'details': {'exists': bool(str(path)) and path.exists(), 'path': str(path) if str(path) else ''},
        },
    }


@app.get('/initializer/status')
def initializer_status():
    path = workbook_path()
    exists = bool(str(path)) and path.exists()
    checks = [
        {'id': 'worker', 'label': 'Local Diehl worker', 'status': 'ok', 'detail': f'Running v{VERSION} on 127.0.0.1:{PORT}'},
        {'id': 'excel', 'label': 'Selected existing workbook', 'status': 'ok' if exists else 'missing', 'detail': str(path) if exists else 'Choose your existing workbook.'},
        {'id': 'dtna', 'label': 'DTNA automation', 'status': 'ok' if DTNA_SCRIPT.exists() else 'missing', 'detail': str(DTNA_SCRIPT)},
        {'id': 'browser', 'label': 'Persistent DTNA browser profile', 'status': 'ok', 'detail': str(PROFILE)},
    ]
    return {'ready': exists and DTNA_SCRIPT.exists(), 'checks': checks, 'summary': 'Local worker ready' if exists else 'Workbook selection required'}


@app.post('/workbook/select')
def workbook_select():
    script = ROOT / 'configure_workbook.py'
    if not script.exists(): raise HTTPException(500, 'Workbook selector is not installed.')
    flags = getattr(subprocess, 'CREATE_NEW_CONSOLE', 0)
    subprocess.Popen([str(ROOT / '.venv' / 'Scripts' / 'python.exe'), str(script)], cwd=str(ROOT), creationflags=flags)
    return {'ok': True}


def launch_dtna(args: list[str]) -> None:
    if not DTNA_SCRIPT.exists(): raise HTTPException(500, 'DTNA automation is not installed.')
    flags = getattr(subprocess, 'CREATE_NEW_CONSOLE', 0)
    subprocess.Popen([str(ROOT / '.venv' / 'Scripts' / 'python.exe'), str(DTNA_SCRIPT), *args], cwd=str(ROOT), creationflags=flags)


@app.get('/dtna/status')
def dtna_status():
    return {'ready': DTNA_SCRIPT.exists(), 'profile': str(PROFILE)}


@app.post('/dtna/open')
def dtna_open():
    launch_dtna(['--login-only'])
    return {'ok': True, 'message': 'DTNA login browser opened.'}


@app.post('/dtna/sync')
def dtna_sync():
    launch_dtna([])
    return {'ok': True, 'message': 'DTNA sync started.'}


@app.post('/batches')
def create_batch(body: BatchIn):
    vins = clean_vins(body.vins)
    if not vins: raise HTTPException(400, 'Enter at least one valid 17-character VIN.')
    path = workbook_path()
    if not path.exists(): raise HTTPException(409, 'The configured workbook cannot be found. Choose it again on the Initializer page.')
    batch_id = str(uuid.uuid4())
    options = {'workers': 1, 'batchSize': max(1, body.batchSize)}
    c = conn()
    c.execute('insert into batches(id,status,total_vins,lookup_mode,options,created_at) values(?,?,?,?,?,?)', (batch_id, 'queued', len(vins), body.lookupMode, json.dumps(options), now()))
    for i, vin in enumerate(vins):
        c.execute('insert into items(id,batch_id,vin,queue_position,status,attempts) values(?,?,?,?,?,0)', (str(uuid.uuid4()), batch_id, vin, i, 'queued'))
    c.commit(); c.close()
    return {'batchId': batch_id, 'total': len(vins)}


@app.get('/batches/resumable')
def resumable():
    c = conn(); row = c.execute("select * from batches where status in ('queued','running','paused') order by created_at desc limit 1").fetchone(); c.close()
    return {'batch': dict(row) if row else None}


@app.get('/batches/{batch_id}')
def get_batch(batch_id: str):
    c = conn(); batch = c.execute('select * from batches where id=?', (batch_id,)).fetchone(); items = c.execute('select * from items where batch_id=? order by queue_position', (batch_id,)).fetchall(); c.close()
    if not batch: raise HTTPException(404, 'Batch not found.')
    out = []
    for item in items:
        d = dict(item); d['result'] = json.loads(d['result']) if d.get('result') else {}; out.append(d)
    return {'batch': dict(batch), 'items': out}


@app.post('/batches/{batch_id}/retry')
def retry_batch(batch_id: str):
    c = conn(); c.execute("update items set status='retry', error_message=null, completed_at=null where batch_id=? and status='error'", (batch_id,)); c.execute("update batches set status='queued', completed_at=null where id=?", (batch_id,)); c.commit(); c.close(); return {'ok': True}


@app.post('/batches/{batch_id}/cancel')
def cancel_batch(batch_id: str):
    c = conn(); c.execute("update batches set status='cancelled', completed_at=? where id=?", (now(), batch_id)); c.execute("update items set status='cancelled' where batch_id=? and status in ('queued','retry')", (batch_id,)); c.commit(); c.close(); return {'ok': True}


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host='127.0.0.1', port=PORT, log_level='warning')
