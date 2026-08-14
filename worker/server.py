import json
import os
import re
import socket
import sqlite3
import subprocess
import threading
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / 'config.json'
DB_PATH = ROOT / 'worker_state.db'
PROFILE_ROOT = ROOT / 'browser_profiles'
DTNA_SCRIPT = ROOT / 'dtna_login_and_sync.py'
PROFILE_ROOT.mkdir(exist_ok=True)

WORKER_VERSION = '3.2'


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def load_config() -> dict[str, Any]:
    if not CONFIG_PATH.exists():
        return {'masterWorkbook': '', 'port': 8765, 'vinLookupCommand': ''}
    try:
        return json.loads(CONFIG_PATH.read_text(encoding='utf-8'))
    except Exception:
        return {'masterWorkbook': '', 'port': 8765, 'vinLookupCommand': ''}


def save_config(config: dict[str, Any]) -> None:
    CONFIG_PATH.write_text(json.dumps(config, indent=2), encoding='utf-8')


def workbook_path() -> Path:
    return Path(os.path.expandvars(str(load_config().get('masterWorkbook', '')))).expanduser()


def port() -> int:
    try:
        return int(load_config().get('port', 8765))
    except Exception:
        return 8765


app = FastAPI(title='Diehl Local VIN Worker', version=WORKER_VERSION)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        'https://diehl-vin-platform.vercel.app',
        'http://localhost:3000',
        'http://127.0.0.1:3000',
    ],
    allow_credentials=False,
    allow_methods=['*'],
    allow_headers=['*'],
)


@app.middleware('http')
async def private_network_access(request: Request, call_next):
    response = await call_next(request)
    if request.headers.get('access-control-request-private-network') == 'true':
        response.headers['Access-Control-Allow-Private-Network'] = 'true'
    response.headers['Cache-Control'] = 'no-store'
    return response


excel_lock = threading.RLock()
stop_event = threading.Event()


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=30)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = db()
    conn.executescript(
        '''
        create table if not exists batches(
          id text primary key,
          status text not null,
          total_vins integer not null,
          lookup_mode text not null,
          options text not null,
          created_at text not null,
          started_at text,
          completed_at text
        );
        create table if not exists items(
          id text primary key,
          batch_id text not null,
          vin text not null,
          queue_position integer not null,
          status text not null,
          attempts integer not null default 0,
          result text,
          error_message text,
          started_at text,
          completed_at text
        );
        create index if not exists items_batch_idx on items(batch_id, queue_position);
        '''
    )
    conn.commit()
    conn.close()


init_db()


def proc_running(*names: str) -> bool:
    try:
        import psutil
        for proc in psutil.process_iter(['name']):
            name = (proc.info.get('name') or '').lower()
            if any(x.lower() in name for x in names):
                return True
    except Exception:
        pass
    return False


def workbook_status() -> dict[str, Any]:
    path = workbook_path()
    exists = bool(str(path)) and path.exists()
    info = {
        'exists': exists,
        'path': str(path) if str(path) else '',
        'name': path.name if str(path) else '',
        'size': path.stat().st_size if exists else 0,
        'modified': datetime.fromtimestamp(path.stat().st_mtime).isoformat() if exists else None,
        'extension': path.suffix.lower() if str(path) else '',
    }
    if exists:
        try:
            read_master(set(), probe_only=True)
            info['accessible'] = True
            info['accessError'] = ''
        except Exception as exc:
            info['accessible'] = False
            info['accessError'] = str(exc)
    else:
        info['accessible'] = False
        info['accessError'] = 'Workbook does not exist.'
    return info


def serializable(value: Any) -> Any:
    return value.isoformat() if hasattr(value, 'isoformat') else value


def _find_open_workbook(excel: Any, path: Path):
    target = os.path.normcase(os.path.abspath(str(path)))
    try:
        for candidate in excel.Workbooks:
            try:
                if os.path.normcase(os.path.abspath(str(candidate.FullName))) == target:
                    return candidate
            except Exception:
                pass
    except Exception:
        pass
    return None


def _read_master_com(path: Path, vins: set[str], probe_only: bool = False) -> dict[str, dict[str, Any]]:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    wb = None
    opened_here = False
    created_excel = False
    try:
        try:
            excel = win32com.client.GetActiveObject('Excel.Application')
        except Exception:
            excel = win32com.client.DispatchEx('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            created_excel = True

        wb = _find_open_workbook(excel, path)
        if wb is None:
            # Excel can usually open a OneDrive workbook read-only even when the
            # file is currently open/sync-locked and direct ZipFile access fails.
            wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=True, IgnoreReadOnlyRecommended=True)
            opened_here = True

        try:
            ws = wb.Worksheets('VIN Data')
        except Exception:
            ws = wb.Worksheets(1)

        used_cols = max(1, int(ws.UsedRange.Columns.Count))
        headers: list[str] = []
        for col in range(1, used_cols + 1):
            headers.append(str(ws.Cells(1, col).Value or '').strip())
        if 'VIN' not in headers:
            raise RuntimeError('Selected workbook must contain a VIN column in row 1.')
        if probe_only:
            return {}

        vin_col = headers.index('VIN') + 1
        last_row = int(ws.Cells(ws.Rows.Count, vin_col).End(-4162).Row)
        result: dict[str, dict[str, Any]] = {}
        for row_num in range(2, max(2, last_row) + 1):
            vin = str(ws.Cells(row_num, vin_col).Value or '').strip().upper()
            if vin and vin in vins:
                row_data: dict[str, Any] = {}
                for col, header in enumerate(headers, start=1):
                    if not header:
                        continue
                    value = ws.Cells(row_num, col).Value
                    if value not in (None, ''):
                        row_data[header] = serializable(value)
                result[vin] = row_data
                if len(result) == len(vins):
                    break
        return result
    finally:
        try:
            if wb is not None and opened_here:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None and created_excel:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def _read_master_openpyxl(path: Path, vins: set[str], probe_only: bool = False) -> dict[str, dict[str, Any]]:
    last_exc: Exception | None = None
    for attempt in range(3):
        wb = None
        try:
            wb = load_workbook(path, read_only=True, data_only=True, keep_vba=path.suffix.lower() == '.xlsm')
            ws = wb['VIN Data'] if 'VIN Data' in wb.sheetnames else wb.active
            rows = ws.iter_rows(values_only=True)
            headers = [str(x or '').strip() for x in next(rows)]
            if 'VIN' not in headers:
                raise RuntimeError('Selected workbook must contain a VIN column in row 1.')
            if probe_only:
                return {}
            idx = {header: i for i, header in enumerate(headers) if header}
            vin_index = idx['VIN']
            result: dict[str, dict[str, Any]] = {}
            for row in rows:
                vin = str(row[vin_index] or '').strip().upper()
                if vin in vins:
                    result[vin] = {
                        headers[i]: serializable(row[i])
                        for i in range(min(len(headers), len(row)))
                        if headers[i] and row[i] not in (None, '')
                    }
                    if len(result) == len(vins):
                        break
            return result
        except Exception as exc:
            last_exc = exc
            time.sleep(0.35 * (attempt + 1))
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass
    raise RuntimeError(f'Excel workbook is temporarily unavailable: {last_exc}')


def read_master(vins: set[str], probe_only: bool = False) -> dict[str, dict[str, Any]]:
    path = workbook_path()
    if not path.exists() or path.suffix.lower() not in {'.xlsx', '.xlsm'}:
        if probe_only:
            raise RuntimeError('The selected Excel workbook does not exist.')
        return {}

    with excel_lock:
        com_error = None
        try:
            return _read_master_com(path, vins, probe_only=probe_only)
        except Exception as exc:
            com_error = exc
        try:
            return _read_master_openpyxl(path, vins, probe_only=probe_only)
        except Exception as file_exc:
            raise RuntimeError(
                'Cannot access the selected Excel workbook right now. '
                'If it is syncing in OneDrive, wait for sync to finish or open it in Excel and try again. '
                f'Excel error: {com_error}; file error: {file_exc}'
            ) from file_exc


def normalize(vin: str, row: dict[str, Any]) -> dict[str, Any]:
    def get(*names: str):
        for name in names:
            if row.get(name) not in (None, ''):
                return row[name]
        return None
    return {
        'vin': vin,
        'verificationStatus': get('Verification Status'),
        'inServiceStatus': get('In-Service Status'),
        'inServiceDate': get('In-Service Date'),
        'mileage': get('Mileage'),
        'customerResult': get('Customer Result'),
        'customerName': get('Customer Name'),
        'registeredCustomerName': get('Registered Customer Name'),
        'registeredCustomerAccount': get('Registered Customer Account'),
        'orderedCustomerName': get('Ordered Customer Name'),
        'raw': row,
    }


def excel_com_write(vin: str, result: dict[str, Any]) -> None:
    path = workbook_path()
    if not path.exists():
        raise RuntimeError('The selected Excel workbook no longer exists.')
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError('Microsoft Excel integration is not installed. Re-run the initializer.') from exc

    mapping = {
        'verificationStatus': 'Verification Status',
        'inServiceStatus': 'In-Service Status',
        'inServiceDate': 'In-Service Date',
        'mileage': 'Mileage',
        'customerResult': 'Customer Result',
        'customerName': 'Customer Name',
        'registeredCustomerName': 'Registered Customer Name',
        'registeredCustomerAccount': 'Registered Customer Account',
        'orderedCustomerName': 'Ordered Customer Name',
    }

    with excel_lock:
        pythoncom.CoInitialize()
        excel = None
        wb = None
        opened_here = False
        created_excel = False
        try:
            try:
                excel = win32com.client.GetActiveObject('Excel.Application')
            except Exception:
                excel = win32com.client.DispatchEx('Excel.Application')
                excel.Visible = False
                excel.DisplayAlerts = False
                created_excel = True

            wb = _find_open_workbook(excel, path)
            if wb is None:
                wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=False, IgnoreReadOnlyRecommended=True)
                opened_here = True

            if bool(getattr(wb, 'ReadOnly', False)):
                raise RuntimeError(
                    'The workbook is open read-only or locked by OneDrive/another Excel session. '
                    'Close duplicate copies and allow OneDrive to finish syncing, then retry.'
                )

            try:
                ws = wb.Worksheets('VIN Data')
            except Exception:
                ws = wb.Worksheets(1)

            used_cols = max(1, int(ws.UsedRange.Columns.Count))
            headers: dict[str, int] = {}
            for col in range(1, used_cols + 1):
                value = str(ws.Cells(1, col).Value or '').strip()
                if value:
                    headers[value] = col
            if 'VIN' not in headers:
                raise RuntimeError('Selected workbook must contain a VIN column in row 1.')

            for heading in mapping.values():
                if heading not in headers:
                    used_cols += 1
                    ws.Cells(1, used_cols).Value = heading
                    headers[heading] = used_cols

            vin_col = headers['VIN']
            last_row = int(ws.Cells(ws.Rows.Count, vin_col).End(-4162).Row)
            row_num = None
            for row in range(2, max(2, last_row) + 1):
                if str(ws.Cells(row, vin_col).Value or '').strip().upper() == vin:
                    row_num = row
                    break

            if row_num is None:
                row_num = max(2, last_row + 1)
                if row_num > 2:
                    ws.Rows(row_num - 1).Copy()
                    ws.Rows(row_num).PasteSpecial(Paste=-4122)
                    ws.Rows(row_num).PasteSpecial(Paste=6)
                    excel.CutCopyMode = False
                ws.Cells(row_num, vin_col).Value = vin

            for key, heading in mapping.items():
                value = result.get(key)
                if value not in (None, ''):
                    ws.Cells(row_num, headers[heading]).Value = value

            wb.Save()
        finally:
            try:
                if wb is not None and opened_here:
                    wb.Close(SaveChanges=True)
            except Exception:
                pass
            try:
                if excel is not None and created_excel:
                    excel.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()


def run_lookup_command(vins: list[str], slot: int) -> dict[str, Any]:
    config = load_config()
    command = str(config.get('vinLookupCommand') or '').strip()
    if not command:
        return {}
    result_file = ROOT / f'results-slot-{slot}.json'
    result_file.unlink(missing_ok=True)
    env = os.environ.copy()
    env['DIEHL_VINS'] = '\n'.join(vins)
    env['DIEHL_RESULT_FILE'] = str(result_file)
    env['DIEHL_WORKER_SLOT'] = str(slot)
    env['DIEHL_BROWSER_PROFILE'] = str(PROFILE_ROOT / f'worker-{slot}')
    completed = subprocess.run(command, shell=True, cwd=str(ROOT), env=env, check=False)
    if completed.returncode != 0 or not result_file.exists():
        return {}
    try:
        raw = json.loads(result_file.read_text(encoding='utf-8'))
        if isinstance(raw, dict):
            return raw
        if isinstance(raw, list):
            return {str(x.get('vin', '')).upper(): x for x in raw if isinstance(x, dict) and x.get('vin')}
    except Exception:
        pass
    return {}


def process_item(item: dict[str, Any], slot: int, existing_row: dict[str, Any] | None = None) -> tuple[dict[str, Any] | None, str | None]:
    vin = item['vin']
    try:
        if existing_row:
            return normalize(vin, existing_row), None

        looked_up = run_lookup_command([vin], slot)
        result = looked_up.get(vin) if isinstance(looked_up, dict) else None
        if result:
            result.setdefault('vin', vin)
            excel_com_write(vin, result)
            return result, None
        return None, 'VIN is not already in the selected workbook and the DTNA lookup did not return a result.'
    except Exception as exc:
        return None, str(exc)


def scheduler() -> None:
    while not stop_event.is_set():
        try:
            conn = db()
            batch = conn.execute("select * from batches where status in ('queued','running') order by created_at limit 1").fetchone()
            if not batch:
                conn.close()
                time.sleep(.8)
                continue

            options = json.loads(batch['options'] or '{}')
            workers = max(1, min(8, int(options.get('workers', 1))))
            conn.execute("update batches set status='running', started_at=coalesce(started_at, ?) where id=?", (now(), batch['id']))
            pending = conn.execute(
                "select * from items where batch_id=? and status in ('queued','retry') order by queue_position limit ?",
                (batch['id'], workers),
            ).fetchall()

            if not pending:
                remaining = conn.execute(
                    "select count(*) as n from items where batch_id=? and status in ('queued','retry','running')",
                    (batch['id'],),
                ).fetchone()['n']
                if remaining == 0:
                    conn.execute("update batches set status='complete', completed_at=? where id=?", (now(), batch['id']))
                conn.commit()
                conn.close()
                time.sleep(.4)
                continue

            # Read the workbook ONCE for this group instead of every worker thread
            # reopening the same OneDrive file independently.
            pending_vins = {str(item['vin']).upper() for item in pending}
            try:
                existing_rows = read_master(pending_vins)
                workbook_error = None
            except Exception as exc:
                existing_rows = {}
                workbook_error = str(exc)

            for item in pending:
                conn.execute(
                    "update items set status='running', started_at=?, attempts=attempts+1 where id=?",
                    (now(), item['id']),
                )
            conn.commit()
            conn.close()

            if workbook_error:
                c = db()
                for item in pending:
                    c.execute(
                        "update items set status='error', error_message=?, completed_at=? where id=?",
                        (workbook_error, now(), item['id']),
                    )
                c.commit()
                c.close()
                time.sleep(1)
                continue

            threads = []

            def run_one(item_row: sqlite3.Row, slot: int):
                c = None
                try:
                    vin = str(item_row['vin']).upper()
                    result, error = process_item(dict(item_row), slot, existing_rows.get(vin))
                    c = db()
                    if result:
                        c.execute(
                            "update items set status='complete', result=?, error_message=null, completed_at=? where id=?",
                            (json.dumps(result, default=str), now(), item_row['id']),
                        )
                    else:
                        c.execute(
                            "update items set status='error', error_message=?, completed_at=? where id=?",
                            (error or 'Unknown worker error', now(), item_row['id']),
                        )
                    c.commit()
                except Exception as exc:
                    try:
                        if c is None:
                            c = db()
                        c.execute(
                            "update items set status='error', error_message=?, completed_at=? where id=?",
                            (str(exc), now(), item_row['id']),
                        )
                        c.commit()
                    except Exception:
                        pass
                finally:
                    try:
                        if c is not None:
                            c.close()
                    except Exception:
                        pass

            for index, item in enumerate(pending):
                thread = threading.Thread(target=run_one, args=(item, index + 1), daemon=True)
                thread.start()
                threads.append(thread)
            for thread in threads:
                thread.join()
            time.sleep(float(options.get('batchPause', .5)))
        except Exception as exc:
            print('scheduler error:', exc, flush=True)
            time.sleep(2)


threading.Thread(target=scheduler, daemon=True).start()


class BatchIn(BaseModel):
    vins: Any
    lookupMode: str = 'in_service_customer'
    workers: int = 1
    batchSize: int = 100
    retryRounds: int = 3
    batchPause: float = .5
    retryPause: float = 3


def clean_vins(value: Any) -> list[str]:
    raw = '\n'.join(value) if isinstance(value, list) else str(value or '')
    return list(dict.fromkeys(
        token.upper()
        for token in re.split(r'[\s,;]+', raw)
        if re.fullmatch(r'[A-HJ-NPR-Z0-9]{17}', token.upper())
    ))


def launch_dtna(extra_args: list[str] | None = None) -> None:
    if not DTNA_SCRIPT.exists():
        raise HTTPException(500, 'DTNA automation is not installed. Re-run the initializer.')
    flags = getattr(subprocess, 'CREATE_NEW_CONSOLE', 0)
    command = [str(ROOT / '.venv' / 'Scripts' / 'python.exe'), str(DTNA_SCRIPT)]
    if extra_args:
        command.extend(extra_args)
    subprocess.Popen(command, cwd=str(ROOT), creationflags=flags)


@app.get('/health')
def health():
    wb = workbook_status()
    return {
        'ok': True,
        'version': WORKER_VERSION,
        'worker': {
            'worker_id': socket.gethostname(),
            'hostname': socket.gethostname(),
            'dtna_status': 'browser open' if proc_running('msedge', 'chrome') else 'ready',
            'master_workbook': wb['path'],
            'last_seen': now(),
            'details': wb,
        },
    }


@app.get('/initializer/status')
def initializer_status():
    wb = workbook_status()
    excel_status = 'ok' if wb.get('accessible') else ('missing' if not wb['exists'] else 'warning')
    checks = [
        {'id': 'worker', 'label': 'Local Diehl worker', 'status': 'ok', 'detail': f'Running v{WORKER_VERSION} on 127.0.0.1:{port()}'},
        {'id': 'excel', 'label': 'Selected existing workbook', 'status': excel_status, 'detail': wb.get('accessError') or wb['path'] or 'No workbook selected'},
        {'id': 'dtna', 'label': 'DTNA automation', 'status': 'ok' if DTNA_SCRIPT.exists() else 'missing', 'detail': str(DTNA_SCRIPT)},
        {'id': 'browser', 'label': 'Persistent DTNA browser', 'status': 'ok' if PROFILE_ROOT.exists() else 'warning', 'detail': str(PROFILE_ROOT)},
        {'id': 'excel-com', 'label': 'Microsoft Excel desktop', 'status': 'ok' if proc_running('excel') else 'warning', 'detail': 'Excel will be opened automatically when workbook access is required.'},
    ]
    missing = [x for x in checks if x['status'] == 'missing']
    return {'ready': not missing and bool(wb.get('accessible')), 'checks': checks, 'summary': f"{len(checks)-len(missing)} of {len(checks)} checks available", 'workbook': wb}


@app.post('/workbook/select')
def select_workbook():
    script = ROOT / 'configure_workbook.py'
    if not script.exists():
        raise HTTPException(500, 'Workbook selector is not installed.')
    flags = getattr(subprocess, 'CREATE_NEW_CONSOLE', 0)
    subprocess.Popen([str(ROOT / '.venv' / 'Scripts' / 'python.exe'), str(script)], cwd=str(ROOT), creationflags=flags)
    return {'ok': True, 'message': 'Workbook selector opened on this computer.'}


@app.get('/dtna/status')
def dtna_status():
    return {
        'ready': DTNA_SCRIPT.exists(),
        'browserOpen': proc_running('msedge', 'chrome'),
        'profile': str(Path(os.environ.get('LOCALAPPDATA', ROOT)) / 'DiehlDTNAManual' / 'browser_profile'),
    }


@app.post('/dtna/open')
def dtna_open():
    launch_dtna(['--login-only'])
    return {'ok': True, 'message': 'DTNA login browser opened on this computer.'}


@app.post('/dtna/sync')
def dtna_sync():
    launch_dtna([])
    return {'ok': True, 'message': 'DTNA Sales Order + AUTO VIN sync started locally.'}


@app.post('/batches')
def create_batch(body: BatchIn):
    vins = clean_vins(body.vins)
    if not vins:
        raise HTTPException(400, 'Enter at least one valid 17-character VIN.')

    wb = workbook_status()
    if not wb['exists']:
        raise HTTPException(409, 'The configured Excel workbook cannot be found. Choose the workbook again on the Initializer page.')
    if not wb.get('accessible'):
        raise HTTPException(409, wb.get('accessError') or 'The configured Excel workbook is not accessible right now.')

    batch_id = str(uuid.uuid4())
    options = {
        'workers': max(1, min(8, body.workers)),
        'batchSize': max(1, body.batchSize),
        'retryRounds': max(0, body.retryRounds),
        'batchPause': max(0, body.batchPause),
        'retryPause': max(0, body.retryPause),
    }
    conn = db()
    conn.execute(
        'insert into batches(id,status,total_vins,lookup_mode,options,created_at) values(?,?,?,?,?,?)',
        (batch_id, 'queued', len(vins), body.lookupMode, json.dumps(options), now()),
    )
    for index, vin in enumerate(vins):
        conn.execute(
            'insert into items(id,batch_id,vin,queue_position,status,attempts) values(?,?,?,?,?,0)',
            (str(uuid.uuid4()), batch_id, vin, index, 'queued'),
        )
    conn.commit()
    conn.close()
    return {'batchId': batch_id, 'total': len(vins)}


@app.get('/batches/resumable')
def resumable():
    conn = db()
    row = conn.execute("select * from batches where status in ('queued','running','paused') order by created_at desc limit 1").fetchone()
    conn.close()
    return {'batch': dict(row) if row else None}


@app.get('/batches/{batch_id}')
def get_batch(batch_id: str):
    conn = db()
    batch = conn.execute('select * from batches where id=?', (batch_id,)).fetchone()
    items = conn.execute('select * from items where batch_id=? order by queue_position', (batch_id,)).fetchall()
    conn.close()
    if not batch:
        raise HTTPException(404, 'Batch not found.')
    output = []
    for item in items:
        data = dict(item)
        data['result'] = json.loads(data['result']) if data.get('result') else {}
        output.append(data)
    return {'batch': dict(batch), 'items': output}


@app.post('/batches/{batch_id}/retry')
def retry_batch(batch_id: str):
    conn = db()
    conn.execute("update items set status='retry', error_message=null, completed_at=null where batch_id=? and status='error'", (batch_id,))
    conn.execute("update batches set status='queued', completed_at=null where id=?", (batch_id,))
    conn.commit()
    conn.close()
    return {'ok': True}


@app.post('/batches/{batch_id}/cancel')
def cancel_batch(batch_id: str):
    conn = db()
    conn.execute("update batches set status='cancelled', completed_at=? where id=?", (now(), batch_id))
    conn.execute("update items set status='cancelled' where batch_id=? and status in ('queued','retry')", (batch_id,))
    conn.commit()
    conn.close()
    return {'ok': True}


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host='127.0.0.1', port=port(), log_level='warning')
