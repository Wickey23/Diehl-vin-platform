from __future__ import annotations

import json
import os
import socket
import threading
import time
import urllib.request
from pathlib import Path
from typing import Any

import psutil
from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
CONFIG = ROOT / 'config.json'
PORT = 8766
ALLOWED_SHEETS = ('DTNA', 'VIN In-Service')
DIEHL_PRODUCTS = {'DiehlVINWorker', 'DiehlVINDatabase'}

app = FastAPI(title='Diehl VIN Database Viewer', version='1.1')
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        'https://diehl-vin-platform.vercel.app',
        'http://localhost:3000',
        'http://127.0.0.1:3000',
    ],
    allow_credentials=False,
    allow_methods=['GET', 'POST', 'OPTIONS'],
    allow_headers=['*'],
)


@app.middleware('http')
async def private_network_access(request: Request, call_next):
    response = await call_next(request)
    if request.headers.get('access-control-request-private-network') == 'true':
        response.headers['Access-Control-Allow-Private-Network'] = 'true'
    response.headers['Cache-Control'] = 'no-store'
    return response


def config() -> dict[str, Any]:
    try:
        return json.loads(CONFIG.read_text(encoding='utf-8')) if CONFIG.exists() else {}
    except Exception:
        return {}


def workbook_path() -> Path:
    value = str(config().get('masterWorkbook') or '').strip()
    return Path(os.path.expandvars(value)).expanduser()


def serialize(value: Any) -> Any:
    if value is None:
        return ''
    if hasattr(value, 'isoformat'):
        try:
            return value.isoformat()
        except Exception:
            pass
    return value if isinstance(value, (str, int, float, bool)) else str(value)


def _rows_to_payload(sheet_name: str, path: Path, headers: list[str], rows_iter, limit: int) -> dict[str, Any]:
    while headers and not headers[-1]:
        headers.pop()
    data: list[dict[str, Any]] = []
    total = 0
    for raw in rows_iter:
        raw_values = list(raw)
        if not any(v not in (None, '') for v in raw_values):
            continue
        total += 1
        if len(data) < limit:
            item: dict[str, Any] = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                value = raw_values[i] if i < len(raw_values) else None
                item[header] = serialize(value)
            data.append(item)
    return {
        'sheet': sheet_name,
        'workbook': str(path),
        'headers': headers,
        'rows': data,
        'rowCount': total,
        'returned': len(data),
        'exists': True,
        'modified': path.stat().st_mtime,
    }


def read_sheet_openpyxl(path: Path, sheet_name: str, limit: int) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=path.suffix.lower() == '.xlsm')
    try:
        if sheet_name not in wb.sheetnames:
            return {'sheet': sheet_name, 'workbook': str(path), 'headers': [], 'rows': [], 'rowCount': 0, 'exists': False}
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            raw_headers = []
        headers = [str(x or '').strip() for x in raw_headers]
        return _rows_to_payload(sheet_name, path, headers, rows, limit)
    finally:
        wb.close()


def read_sheet_com(path: Path, sheet_name: str, limit: int) -> dict[str, Any]:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = wb = None
    opened_here = created_excel = False
    try:
        target = os.path.normcase(os.path.abspath(str(path)))
        try:
            excel = win32com.client.GetActiveObject('Excel.Application')
            for candidate in excel.Workbooks:
                try:
                    if os.path.normcase(os.path.abspath(str(candidate.FullName))) == target:
                        wb = candidate
                        break
                except Exception:
                    pass
        except Exception:
            excel = None

        if wb is None:
            if excel is None:
                excel = win32com.client.DispatchEx('Excel.Application')
                excel.Visible = False
                excel.DisplayAlerts = False
                created_excel = True
            wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=True, IgnoreReadOnlyRecommended=True, AddToMru=False)
            opened_here = True

        try:
            ws = wb.Worksheets(sheet_name)
        except Exception:
            return {'sheet': sheet_name, 'workbook': str(path), 'headers': [], 'rows': [], 'rowCount': 0, 'exists': False}

        used = ws.UsedRange
        row_count = max(1, int(used.Rows.Count))
        col_count = max(1, int(used.Columns.Count))
        headers = [str(ws.Cells(1, c).Value or '').strip() for c in range(1, col_count + 1)]

        def values():
            for r in range(2, row_count + 1):
                yield [ws.Cells(r, c).Value for c in range(1, col_count + 1)]

        return _rows_to_payload(sheet_name, path, headers, values(), limit)
    finally:
        try:
            if wb is not None and opened_here:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None and created_excel:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def read_sheet(sheet_name: str, limit: int) -> dict[str, Any]:
    path = workbook_path()
    if not str(path) or not path.exists():
        raise RuntimeError('Shared Excel database was not found on this computer.')

    errors: list[str] = []
    for attempt in range(3):
        try:
            return read_sheet_openpyxl(path, sheet_name, limit)
        except Exception as exc:
            errors.append(f'file read: {exc}')
            try:
                return read_sheet_com(path, sheet_name, limit)
            except Exception as com_exc:
                errors.append(f'Excel read: {com_exc}')
            if attempt < 2:
                time.sleep(.5)
    raise RuntimeError('Could not read the shared Excel database: ' + ' | '.join(errors[-4:]))


def ping_product(port: int) -> str:
    try:
        with urllib.request.urlopen(f'http://127.0.0.1:{port}/ping', timeout=1) as response:
            data = json.loads(response.read().decode('utf-8', errors='replace'))
        return str(data.get('product') or '')
    except Exception:
        return ''


def listener_pid(port: int) -> int | None:
    try:
        for conn in psutil.net_connections(kind='tcp'):
            if conn.status == psutil.CONN_LISTEN and conn.laddr and conn.laddr.port == port and conn.pid:
                return int(conn.pid)
    except Exception:
        pass
    return None


def verified_diehl_process(port: int, pid: int) -> bool:
    product = ping_product(port)
    if product not in DIEHL_PRODUCTS:
        return False
    try:
        proc = psutil.Process(pid)
        text = ' '.join(proc.cmdline()).lower()
    except Exception:
        return False
    if port == 8765:
        return 'service_v4.py' in text or 'diehlvinworker' in text
    if port == 8766:
        return 'database_service.py' in text or pid == os.getpid()
    return False


def _stop_main_worker() -> None:
    pid = listener_pid(8765)
    if pid and verified_diehl_process(8765, pid):
        try:
            psutil.Process(pid).terminate()
            try:
                psutil.Process(pid).wait(timeout=3)
            except Exception:
                if psutil.pid_exists(pid):
                    psutil.Process(pid).kill()
        except Exception:
            pass


def _exit_database_service() -> None:
    time.sleep(.8)
    os._exit(0)


@app.get('/ping')
def ping():
    return {'ok': True, 'product': 'DiehlVINDatabase', 'version': '1.1', 'hostname': socket.gethostname()}


@app.get('/database/sheets')
def sheets():
    path = workbook_path()
    return {'ok': True, 'workbook': str(path) if str(path) else '', 'sheets': list(ALLOWED_SHEETS)}


@app.get('/database/{sheet_name}')
def database_sheet(sheet_name: str, limit: int = Query(default=2000, ge=1, le=10000)):
    if sheet_name not in ALLOWED_SHEETS:
        raise HTTPException(404, 'Unknown database sheet.')
    try:
        return read_sheet(sheet_name, limit)
    except Exception as exc:
        raise HTTPException(503, str(exc)) from exc


@app.post('/control/stop-all')
def stop_all():
    """Stop only verified Diehl local services. Never touches Excel, Edge, or unrelated processes."""
    threading.Thread(target=_stop_main_worker, daemon=True).start()
    threading.Thread(target=_exit_database_service, daemon=True).start()
    return {'ok': True, 'message': 'Stopping Diehl worker services on ports 8765 and 8766.'}


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host='127.0.0.1', port=PORT, log_level='warning')
