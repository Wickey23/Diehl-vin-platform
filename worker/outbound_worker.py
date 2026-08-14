import json, os, socket, subprocess, sys, time
from pathlib import Path
from datetime import datetime, timezone
import requests
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parent
ENV=ROOT/'worker.env'

def load_env():
    if ENV.exists():
        for line in ENV.read_text(encoding='utf-8').splitlines():
            if '=' in line and not line.lstrip().startswith('#'):
                k,v=line.split('=',1); os.environ[k.strip()]=os.path.expandvars(v.strip())
load_env()
BASE=os.environ.get('VERCEL_BASE_URL','https://diehl-vin-platform.vercel.app').rstrip('/')
SECRET=os.environ.get('DTNA_WORKER_SECRET','').strip()
WORKBOOK=Path(os.environ.get('MASTER_WORKBOOK','')).expanduser()
WORKER_ID=os.environ.get('WORKER_ID',socket.gethostname())
POLL=max(2,int(os.environ.get('POLL_SECONDS','4')))
HEADERS={'content-type':'application/json','x-worker-secret':SECRET}

def now(): return datetime.now(timezone.utc).isoformat()
def api(path,payload=None):
    r=requests.post(BASE+path,headers=HEADERS,json=payload or {},timeout=40)
    r.raise_for_status(); return r.json()
def workbook_status():
    return {'exists':WORKBOOK.exists(),'path':str(WORKBOOK),'name':WORKBOOK.name if WORKBOOK else '','extension':WORKBOOK.suffix.lower() if WORKBOOK else ''}
def heartbeat(dtna='ready',message=''):
    api('/api/worker/heartbeat',{'workerId':WORKER_ID,'hostname':socket.gethostname(),'dtnaStatus':dtna,'masterWorkbook':str(WORKBOOK),'details':{'workbook':workbook_status(),'message':message,'version':'3.0-outbound','excelWriteMode':'COM'}})
def read_existing(vin):
    if not WORKBOOK.exists(): return None
    wb=load_workbook(WORKBOOK,read_only=True,data_only=True,keep_vba=WORKBOOK.suffix.lower()=='.xlsm')
    ws=wb['VIN Data'] if 'VIN Data' in wb.sheetnames else wb.active
    headers=[str(c.value or '').strip() for c in next(ws.iter_rows())]; idx={h:i for i,h in enumerate(headers)}; vi=idx.get('VIN',0)
    for row in ws.iter_rows(values_only=True):
        if str(row[vi] or '').strip().upper()==vin:
            data={headers[i]:row[i] for i in range(min(len(headers),len(row))) if headers[i] and row[i] not in (None,'')}; wb.close(); return data
    wb.close(); return None
def normalize(vin,row):
    def g(*names):
        for n in names:
            if row.get(n) not in (None,''): return row[n]
    return {'vin':vin,'verificationStatus':g('Verification Status'),'inServiceStatus':g('In-Service Status'),'inServiceDate':str(g('In-Service Date') or ''),'mileage':g('Mileage'),'customerResult':g('Customer Result'),'customerName':g('Customer Name'),'registeredCustomerName':g('Registered Customer Name'),'registeredCustomerAccount':g('Registered Customer Account'),'orderedCustomerName':g('Ordered Customer Name')}
def excel_write(vin,result):
    try:
        import win32com.client
    except Exception as e: raise RuntimeError('Microsoft Excel automation is unavailable. Re-run installer.') from e
    xl=win32com.client.DispatchEx('Excel.Application'); xl.Visible=False; xl.DisplayAlerts=False
    wb=None
    try:
        wb=xl.Workbooks.Open(str(WORKBOOK),UpdateLinks=0,ReadOnly=False)
        names=[ws.Name for ws in wb.Worksheets]; ws=wb.Worksheets('VIN Data') if 'VIN Data' in names else wb.Worksheets(1)
        used=ws.UsedRange; last_col=max(1,used.Columns.Count); headers={}
        for c in range(1,last_col+1):
            val=str(ws.Cells(1,c).Value or '').strip()
            if val: headers[val]=c
        required=['VIN','Verification Status','In-Service Status','In-Service Date','Mileage','Customer Result','Customer Name','Registered Customer Name','Registered Customer Account','Ordered Customer Name']
        for h in required:
            if h not in headers:
                last_col+=1; ws.Cells(1,last_col).Value=h; headers[h]=last_col
        last_row=max(1,used.Rows.Count); rownum=None
        for r in range(2,last_row+2):
            if str(ws.Cells(r,headers['VIN']).Value or '').strip().upper()==vin: rownum=r; break
        if rownum is None:
            rownum=last_row+1 if last_row>=2 else 2
            ws.Cells(rownum,headers['VIN']).Value=vin
            if rownum>2:
                src=ws.Rows(rownum-1); dst=ws.Rows(rownum); src.Copy(); dst.PasteSpecial(-4122); ws.Cells(rownum,headers['VIN']).Value=vin
        mapping={'verificationStatus':'Verification Status','inServiceStatus':'In-Service Status','inServiceDate':'In-Service Date','mileage':'Mileage','customerResult':'Customer Result','customerName':'Customer Name','registeredCustomerName':'Registered Customer Name','registeredCustomerAccount':'Registered Customer Account','orderedCustomerName':'Ordered Customer Name'}
        for k,h in mapping.items():
            if result.get(k) not in (None,''): ws.Cells(rownum,headers[h]).Value=result[k]
        wb.Save()
    finally:
        if wb: wb.Close(SaveChanges=True)
        xl.Quit()
def run_lookup(vin,slot):
    existing=read_existing(vin)
    if existing: return normalize(vin,existing)
    cmd=os.environ.get('VIN_LOOKUP_COMMAND','').strip()
    if not cmd: raise RuntimeError('VIN In-Service lookup engine is not configured')
    result_file=ROOT/f'result-{slot}.json'; result_file.unlink(missing_ok=True)
    env=os.environ.copy(); env.update({'DIEHL_VINS':vin,'DIEHL_RESULT_FILE':str(result_file),'DIEHL_WORKER_SLOT':str(slot),'DIEHL_BROWSER_PROFILE':str(ROOT/'browser_profiles'/f'worker-{slot}')})
    subprocess.run(cmd,shell=True,cwd=str(ROOT),env=env,check=False)
    if not result_file.exists(): raise RuntimeError('VIN lookup produced no result')
    raw=json.loads(result_file.read_text(encoding='utf-8')); result=raw.get(vin) if isinstance(raw,dict) and vin in raw else raw
    if isinstance(result,list): result=result[0] if result else None
    if not isinstance(result,dict): raise RuntimeError('VIN lookup result was invalid')
    result['vin']=vin; excel_write(vin,result); return result
def run_dtna_command(command):
    script=ROOT/'dtna_login_and_sync.py'
    if command in ('dtna_open','dtna_sync'):
        subprocess.Popen([sys.executable,str(script)],cwd=str(ROOT),creationflags=getattr(subprocess,'CREATE_NEW_CONSOLE',0)); return {'message':'DTNA browser/sync launched locally'}
    return {'message':'Unknown command'}
def main():
    if not SECRET: raise SystemExit('DTNA_WORKER_SECRET is missing from worker.env')
    if not WORKBOOK: raise SystemExit('MASTER_WORKBOOK is missing from worker.env')
    print('Diehl VIN Worker - outbound mode'); print('Vercel:',BASE); print('Workbook:',WORKBOOK); print('Worker:',WORKER_ID)
    while True:
        try:
            heartbeat()
            cmds=api('/api/worker/commands',{'workerId':WORKER_ID,'limit':3})
            for cmd in cmds.get('commands',[]):
                try: result=run_dtna_command(cmd['command']); api('/api/worker/commands/result',{'id':cmd['id'],'status':'complete','result':result})
                except Exception as e: api('/api/worker/commands/result',{'id':cmd['id'],'status':'error','error':str(e)})
            work=api('/api/worker/claim',{'workerId':WORKER_ID,'limit':8})
            batch=work.get('batch'); items=work.get('items') or []
            for slot,item in enumerate(items,1):
                try:
                    result=run_lookup(item['vin'],slot); api('/api/worker/result',{'itemId':item['id'],'batchId':batch['id'],'status':'complete','result':result,'attempts':int(item.get('attempts') or 0)+1})
                except Exception as e:
                    api('/api/worker/result',{'itemId':item['id'],'batchId':batch['id'],'status':'error','error':str(e),'attempts':int(item.get('attempts') or 0)+1})
        except KeyboardInterrupt: break
        except Exception as e:
            print('Worker connection error:',e)
        time.sleep(POLL)
if __name__=='__main__': main()
